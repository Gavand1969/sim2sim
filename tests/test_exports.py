"""
Tests for Pro-tier Excel + PDF export module.

Coverage:
- Each xlsx builder returns non-empty bytes that openpyxl can re-open.
- Each pdf builder returns bytes starting with b"%PDF".
- End-to-end API test: POST /api/export/queuing/xlsx with X-License-Key: dev.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from main import app
from src.export.excel import build_inventory_xlsx, build_lp_xlsx, build_queuing_xlsx
from src.export.pdf import build_inventory_pdf, build_lp_pdf, build_queuing_pdf

client = TestClient(app, raise_server_exceptions=True)

# ── Sample data fixtures ───────────────────────────────────────────────────────

QUEUING_PARAMS = {
    "model": "MM1",
    "arrival_rate": 4.0,
    "service_rate": 6.0,
    "num_servers": 1,
}

QUEUING_RESULT = {
    "model": "MM1",
    "utilization": 0.6667,
    "L": 2.0,
    "Lq": 1.3333,
    "W": 0.5,
    "Wq": 0.3333,
    "P0": 0.3333,
    "P_wait": 0.6667,
    "prob_dist": [0.3333, 0.2222, 0.1481, 0.0988, 0.0658,
                  0.0439, 0.0293, 0.0195, 0.013, 0.0087,
                  0.0058, 0.0039, 0.0026, 0.0017, 0.0011,
                  0.0008, 0.0005, 0.0003, 0.0002, 0.0002, 0.0001],
    "little_law_check": 1e-10,
    "blocking_prob": None,
    "effective_lam": None,
    "notes": None,
}

INVENTORY_PARAMS = {
    "annual_demand": 10000,
    "ordering_cost": 200,
    "unit_cost": 50,
    "holding_rate": 0.25,
}

INVENTORY_RESULT = {
    "eoq": 400.0,
    "orders_per_year": 25.0,
    "cycle_time_days": 14.6,
    "total_annual_cost": 5000.0,
    "cost_curve_q": [100.0, 200.0, 400.0, 600.0, 800.0],
    "cost_curve_tc": [5800.0, 5200.0, 5000.0, 5133.0, 5350.0],
    "cost_curve_holding": [1250.0, 2500.0, 5000.0, 7500.0, 10000.0],
    "cost_curve_ordering": [20000.0, 10000.0, 5000.0, 3333.0, 2500.0],
}

LP_PARAMS = {
    "objective": "maximize",
    "c_obj": [5.0, 4.0],
    "A_ub": [[6.0, 4.0], [1.0, 2.0]],
    "b_ub": [24.0, 6.0],
    "variable_names": ["x1", "x2"],
    "constraint_names": ["c1", "c2"],
}

LP_RESULT = {
    "status": "optimal",
    "optimal_value": 21.0,
    "variables": {"x1": 3.0, "x2": 1.5},
    "shadow_prices": {"c1": 0.75, "c2": 0.25},
    "reduced_costs": {"x1": 0.0, "x2": 0.0},
    "binding_constraints": ["c1"],
    "slacks": {"c1": 0.0, "c2": 2.0},
    "rhs_ranges": {
        "c1": {"lower": 18.0, "current": 24.0, "upper": 30.0},
        "c2": {"lower": 3.0, "current": 6.0, "upper": 12.0},
    },
    "obj_ranges": {
        "x1": {"lower": 4.0, "current": 5.0, "upper": 8.0},
        "x2": {"lower": 2.5, "current": 4.0, "upper": 5.0},
    },
}


# ── Excel builder tests ────────────────────────────────────────────────────────

class TestQueuingXlsx:
    def test_returns_nonempty_bytes(self):
        data = build_queuing_xlsx(QUEUING_RESULT, QUEUING_PARAMS)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_openpyxl_can_reopen(self):
        data = build_queuing_xlsx(QUEUING_RESULT, QUEUING_PARAMS)
        wb = load_workbook(BytesIO(data))
        assert wb is not None

    def test_correct_sheet_names(self):
        data = build_queuing_xlsx(QUEUING_RESULT, QUEUING_PARAMS)
        wb = load_workbook(BytesIO(data))
        assert set(wb.sheetnames) == {"Summary", "Probability Distribution", "Inputs"}

    def test_summary_has_utilization(self):
        data = build_queuing_xlsx(QUEUING_RESULT, QUEUING_PARAMS)
        wb = load_workbook(BytesIO(data))
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert any("Utilization" in str(lbl) for lbl in labels if lbl)

    def test_prob_dist_sheet_has_21_data_rows(self):
        data = build_queuing_xlsx(QUEUING_RESULT, QUEUING_PARAMS)
        wb = load_workbook(BytesIO(data))
        ws = wb["Probability Distribution"]
        # Row 1 = header, rows 2-22 = data (21 rows)
        data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value is not None]
        assert len(data_rows) == 21


class TestInventoryXlsx:
    def test_returns_nonempty_bytes(self):
        data = build_inventory_xlsx(INVENTORY_RESULT, INVENTORY_PARAMS, "eoq")
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_openpyxl_can_reopen(self):
        data = build_inventory_xlsx(INVENTORY_RESULT, INVENTORY_PARAMS, "eoq")
        wb = load_workbook(BytesIO(data))
        assert wb is not None

    def test_correct_sheet_names(self):
        data = build_inventory_xlsx(INVENTORY_RESULT, INVENTORY_PARAMS, "eoq")
        wb = load_workbook(BytesIO(data))
        assert "Summary" in wb.sheetnames
        assert "Cost Curve" in wb.sheetnames
        assert "Inputs" in wb.sheetnames

    def test_no_list_fields_in_summary(self):
        data = build_inventory_xlsx(INVENTORY_RESULT, INVENTORY_PARAMS, "eoq")
        wb = load_workbook(BytesIO(data))
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert all("cost_curve" not in str(lbl) for lbl in labels if lbl)


class TestLPXlsx:
    def test_returns_nonempty_bytes(self):
        data = build_lp_xlsx(LP_RESULT, LP_PARAMS)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_openpyxl_can_reopen(self):
        data = build_lp_xlsx(LP_RESULT, LP_PARAMS)
        wb = load_workbook(BytesIO(data))
        assert wb is not None

    def test_correct_sheet_names(self):
        data = build_lp_xlsx(LP_RESULT, LP_PARAMS)
        wb = load_workbook(BytesIO(data))
        assert set(wb.sheetnames) == {"Solution", "Constraints", "Objective Ranging"}


# ── PDF builder tests ──────────────────────────────────────────────────────────

class TestQueuingPdf:
    def test_returns_bytes_starting_with_pdf_header(self):
        data = build_queuing_pdf(QUEUING_RESULT, QUEUING_PARAMS)
        assert isinstance(data, bytes)
        assert data[:4] == b"%PDF"

    def test_nonempty(self):
        data = build_queuing_pdf(QUEUING_RESULT, QUEUING_PARAMS)
        assert len(data) > 1000  # PDFs are always larger than this


class TestInventoryPdf:
    def test_returns_bytes_starting_with_pdf_header(self):
        data = build_inventory_pdf(INVENTORY_RESULT, INVENTORY_PARAMS, "eoq")
        assert isinstance(data, bytes)
        assert data[:4] == b"%PDF"

    def test_nonempty(self):
        data = build_inventory_pdf(INVENTORY_RESULT, INVENTORY_PARAMS, "eoq")
        assert len(data) > 1000


class TestLPPdf:
    def test_returns_bytes_starting_with_pdf_header(self):
        data = build_lp_pdf(LP_RESULT, LP_PARAMS)
        assert isinstance(data, bytes)
        assert data[:4] == b"%PDF"

    def test_nonempty(self):
        data = build_lp_pdf(LP_RESULT, LP_PARAMS)
        assert len(data) > 1000


# ── End-to-end API tests ───────────────────────────────────────────────────────

class TestExportAPI:
    def test_queuing_xlsx_endpoint_returns_valid_xlsx(self):
        r = client.post(
            "/api/export/queuing/xlsx",
            json={"params": QUEUING_PARAMS, "result": QUEUING_RESULT},
            headers={"X-License-Key": "dev"},
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert "attachment" in r.headers["content-disposition"]
        assert "sim2sim-queuing.xlsx" in r.headers["content-disposition"]

        # Verify openpyxl can reopen the response body
        wb = load_workbook(BytesIO(r.content))
        assert "Summary" in wb.sheetnames

    def test_queuing_pdf_endpoint_returns_pdf(self):
        r = client.post(
            "/api/export/queuing/pdf",
            json={"params": QUEUING_PARAMS, "result": QUEUING_RESULT},
            headers={"X-License-Key": "dev"},
        )
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    def test_inventory_xlsx_endpoint_returns_valid_xlsx(self):
        r = client.post(
            "/api/export/inventory/xlsx",
            json={
                "params": INVENTORY_PARAMS,
                "result": INVENTORY_RESULT,
                "model_kind": "eoq",
            },
            headers={"X-License-Key": "dev"},
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert "Summary" in wb.sheetnames

    def test_lp_xlsx_endpoint_returns_valid_xlsx(self):
        r = client.post(
            "/api/export/lp/xlsx",
            json={"params": LP_PARAMS, "result": LP_RESULT},
            headers={"X-License-Key": "dev"},
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert "Solution" in wb.sheetnames

    def test_invalid_model_kind_rejected(self):
        r = client.post(
            "/api/export/inventory/xlsx",
            json={
                "params": INVENTORY_PARAMS,
                "result": INVENTORY_RESULT,
                "model_kind": "invalid_model",
            },
            headers={"X-License-Key": "dev"},
        )
        assert r.status_code == 422
