"""
Excel workbook builders for Pro-tier export feature.
Uses openpyxl to produce .xlsx files in-memory.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Shared helpers ─────────────────────────────────────────────────────────────

_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # navy
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_header_row(ws, headers: list[str]) -> None:
    """Write a bold header row in row 1 with navy background."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")


def _freeze_and_autosize(ws, first_col_width: int = 28) -> None:
    """Freeze row 1 and set a reasonable width for column A."""
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = first_col_width


def _write_kv_section(ws, data: dict[str, Any], start_row: int = 2) -> int:
    """Write label/value pairs starting at start_row; return next empty row."""
    row = start_row
    for key, value in data.items():
        if value is None:
            continue
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Queuing workbook ───────────────────────────────────────────────────────────

def build_queuing_xlsx(result: dict, params: dict) -> bytes:
    """Build a 3-sheet workbook for queuing analysis results."""
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_header_row(ws_sum, ["Metric", "Value"])
    _freeze_and_autosize(ws_sum)

    summary_keys = [
        ("model", "Model"),
        ("utilization", "Utilization (ρ)"),
        ("L", "Avg Customers in System (L)"),
        ("Lq", "Avg Customers in Queue (Lq)"),
        ("W", "Avg Time in System (W)"),
        ("Wq", "Avg Time in Queue (Wq)"),
        ("P0", "Probability System Empty (P0)"),
        ("P_wait", "Probability of Wait (P_wait)"),
        ("little_law_check", "Little's Law Check"),
        ("blocking_prob", "Blocking Probability"),
        ("effective_lam", "Effective Arrival Rate"),
        ("notes", "Notes"),
    ]
    row = 2
    for key, label in summary_keys:
        val = result.get(key)
        if val is None:
            continue
        ws_sum.cell(row=row, column=1, value=label)
        ws_sum.cell(row=row, column=2, value=val)
        row += 1

    # ── Sheet 2: Probability Distribution ────────────────────────────────────
    ws_prob = wb.create_sheet("Probability Distribution")
    _write_header_row(ws_prob, ["n", "P(N=n)"])
    _freeze_and_autosize(ws_prob, first_col_width=8)

    prob_dist = result.get("prob_dist", [])
    for n, p in enumerate(prob_dist[:21]):
        ws_prob.cell(row=n + 2, column=1, value=n)
        ws_prob.cell(row=n + 2, column=2, value=p)

    # ── Sheet 3: Inputs ───────────────────────────────────────────────────────
    ws_inp = wb.create_sheet("Inputs")
    _write_header_row(ws_inp, ["Parameter", "Value"])
    _freeze_and_autosize(ws_inp)
    _write_kv_section(ws_inp, params)

    return _to_bytes(wb)


# ── Inventory workbook ─────────────────────────────────────────────────────────

def build_inventory_xlsx(result: dict, params: dict, model_kind: str) -> bytes:
    """Build a 3-sheet workbook for inventory model results."""
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_header_row(ws_sum, ["Metric", "Value"])
    _freeze_and_autosize(ws_sum)

    row = 2
    # Add model kind as first row
    ws_sum.cell(row=row, column=1, value="Model")
    ws_sum.cell(row=row, column=2, value=model_kind)
    row += 1
    for key, value in result.items():
        if value is None:
            continue
        if isinstance(value, list):
            continue  # skip list/curve fields
        ws_sum.cell(row=row, column=1, value=key)
        ws_sum.cell(row=row, column=2, value=value)
        row += 1

    # ── Sheet 2: Cost Curve (if available) ───────────────────────────────────
    has_cost_curve = "cost_curve_q" in result or "profit_curve_q" in result
    if has_cost_curve:
        ws_curve = wb.create_sheet("Cost Curve")

        # Determine column layout based on what's available
        qs = result.get("cost_curve_q", result.get("profit_curve_q", []))
        tcs = result.get("cost_curve_tc", result.get("profit_curve_ep", []))
        holdings = result.get("cost_curve_holding", [])
        orderings = result.get("cost_curve_ordering", [])

        headers = ["Q"]
        if tcs:
            headers.append("TC" if "cost_curve_tc" in result else "Expected Profit")
        if holdings:
            headers.append("Holding")
        if orderings:
            headers.append("Ordering")

        _write_header_row(ws_curve, headers)
        _freeze_and_autosize(ws_curve, first_col_width=12)

        for i, q in enumerate(qs):
            ws_curve.cell(row=i + 2, column=1, value=q)
            col = 2
            if tcs and i < len(tcs):
                ws_curve.cell(row=i + 2, column=col, value=tcs[i])
                col += 1
            if holdings and i < len(holdings):
                ws_curve.cell(row=i + 2, column=col, value=holdings[i])
                col += 1
            if orderings and i < len(orderings):
                ws_curve.cell(row=i + 2, column=col, value=orderings[i])

    # ── Sheet 3: Inputs ───────────────────────────────────────────────────────
    ws_inp = wb.create_sheet("Inputs")
    _write_header_row(ws_inp, ["Parameter", "Value"])
    _freeze_and_autosize(ws_inp)
    _write_kv_section(ws_inp, params)

    return _to_bytes(wb)


# ── LP workbook ────────────────────────────────────────────────────────────────

def build_lp_xlsx(result: dict, params: dict) -> bytes:
    """Build a 3-sheet workbook for LP optimization results."""
    wb = Workbook()

    # ── Sheet 1: Solution ─────────────────────────────────────────────────────
    ws_sol = wb.active
    ws_sol.title = "Solution"
    _write_header_row(ws_sol, ["Item", "Value"])
    _freeze_and_autosize(ws_sol)

    row = 2
    ws_sol.cell(row=row, column=1, value="Status")
    ws_sol.cell(row=row, column=2, value=result.get("status", ""))
    row += 1
    ws_sol.cell(row=row, column=1, value="Optimal Value")
    ws_sol.cell(row=row, column=2, value=result.get("optimal_value"))
    row += 2  # blank separator

    # Variable rows with header
    ws_sol.cell(row=row, column=1, value="Variable")
    ws_sol.cell(row=row, column=2, value="Value")
    ws_sol.cell(row=row, column=3, value="Reduced Cost")
    for col_idx in range(1, 4):
        cell = ws_sol.cell(row=row, column=col_idx)
        cell.font = _BOLD
    row += 1

    variables = result.get("variables", {})
    reduced_costs = result.get("reduced_costs", {})
    for name, value in variables.items():
        ws_sol.cell(row=row, column=1, value=name)
        ws_sol.cell(row=row, column=2, value=value)
        ws_sol.cell(row=row, column=3, value=reduced_costs.get(name))
        row += 1

    # ── Sheet 2: Constraints ──────────────────────────────────────────────────
    ws_con = wb.create_sheet("Constraints")
    _write_header_row(ws_con, ["Name", "Shadow Price", "Slack", "Binding", "RHS Lower", "RHS Current", "RHS Upper"])
    _freeze_and_autosize(ws_con)
    ws_con.column_dimensions["B"].width = 14
    ws_con.column_dimensions["C"].width = 12
    ws_con.column_dimensions["D"].width = 10
    ws_con.column_dimensions["E"].width = 12
    ws_con.column_dimensions["F"].width = 14
    ws_con.column_dimensions["G"].width = 12

    shadow_prices = result.get("shadow_prices", {})
    slacks = result.get("slacks", {})
    binding_constraints = result.get("binding_constraints", [])
    rhs_ranges = result.get("rhs_ranges", {})

    row = 2
    for name in shadow_prices:
        rhs_range = rhs_ranges.get(name, {})
        ws_con.cell(row=row, column=1, value=name)
        ws_con.cell(row=row, column=2, value=shadow_prices.get(name))
        ws_con.cell(row=row, column=3, value=slacks.get(name))
        ws_con.cell(row=row, column=4, value="Y" if name in binding_constraints else "N")
        ws_con.cell(row=row, column=5, value=rhs_range.get("lower"))
        ws_con.cell(row=row, column=6, value=rhs_range.get("current"))
        ws_con.cell(row=row, column=7, value=rhs_range.get("upper"))
        row += 1

    # ── Sheet 3: Objective Ranging ────────────────────────────────────────────
    ws_obj = wb.create_sheet("Objective Ranging")
    _write_header_row(ws_obj, ["Variable", "Obj Lower", "Obj Current", "Obj Upper"])
    _freeze_and_autosize(ws_obj)

    obj_ranges = result.get("obj_ranges", {})
    row = 2
    for var_name, obj_range in obj_ranges.items():
        ws_obj.cell(row=row, column=1, value=var_name)
        ws_obj.cell(row=row, column=2, value=obj_range.get("lower") if isinstance(obj_range, dict) else None)
        ws_obj.cell(row=row, column=3, value=obj_range.get("current") if isinstance(obj_range, dict) else None)
        ws_obj.cell(row=row, column=4, value=obj_range.get("upper") if isinstance(obj_range, dict) else None)
        row += 1

    return _to_bytes(wb)
